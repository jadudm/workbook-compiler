import re

import openpyxl as pxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension, DimensionHolder
from openpyxl.styles import Alignment
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.styles import (
    NamedStyle as PyxlNamedStyle,
    Font as PyxlFont,
    Border as PyxlBorder,
    Side as PyxlSide,
    PatternFill as PyxlPatternFill
)
from colour import Color

from wbc.parser.workbook import Workbook
from wbc.parser.range import Range, LinearRange
from wbc.parser.cell import Cell
from wbc.parser.style import Font
from wbc.parser.exceptions import RenderException


def normalize(s):
    s = re.sub("\\s+", "_", s)
    s = s.lower()
    return s


def add_named_range(opwb, opsh, sh, r, prepend_sheet_name=False):
    # Add as a global named range
    range_start = r.start.as_a1()
    range_end = r.end.as_a1()
    ref = f"{quote_sheetname(opsh.title)}!{absolute_coordinate(range_start + ':' + range_end)}"
    range_name = r.name
    if prepend_sheet_name:
        range_name = normalize(f"{sh.name}_{r.name}")
    else:
        # In case a named range is duplicated on the global table
        # Automatically prepend in that case.
        for k, _ in opwb.defined_names.items():
            if r.name == k:
                range_name = normalize(f"{sh.name}_{r.name}")
    opwb.defined_names[range_name] = DefinedName(range_name, attr_text=ref)


# This asks for a list of locations from the range.
# It will create Cell objects for each location, and
# populate it with its contents. This way, we can easily
# populate cells in the workbook from those Cell objects.
def fill_range_values(wbcwb, opsh, r):
    # A dictionary to build up dimensions of columns
    if isinstance(r, LinearRange):
        r: LinearRange
        if r.header:
            wbc = opsh[r.header.as_a1()]
            wbc.value = f"{r.header.contents.value}"
            if r.header.contents.style:
                wbc.style = r.header.contents.style
        if r.contents:
            for cell in r.locations():
                cell: Cell
                wbc = opsh[cell.as_a1()]
                wbc.value = f"{cell.contents.value}"
                if cell.contents.style:
                    wbc.style = cell.contents.style
                if cell.contents.wrap:
                    wbc.alignment = Alignment(wrap_text=cell.contents.wrap)
                if r.width:
                    opsh.column_dimensions[cell.get_column_as_a1()].width = r.width
                if r.height:
                    opsh.row_dimensions[cell.row].height = r.height

        if r.dynamic:
            m = getattr(wbcwb, "functions")
            fun = getattr(m, r.dynamic)
            for cell in r.locations():
                cell: Cell
                wbc = opsh[cell.as_a1()]
                wbc.value = fun(cell)
                if r.style:
                    wbc.style = r.style
        if r.validation:
            pass
        if r.function1:
            for cell in r.locations():
                cell: Cell
                wbc = opsh[cell.as_a1()]
                wbc.value = r.function1
                if r.style:
                    wbc.style = r.style

    elif isinstance(r, Range):
        r: Range
        # raise RenderException("cannot render Range")
        pass
    else:
        raise RenderException(f"cannot render range of type {type(r)}")


def color2argb(c):
    return c.hex_l.replace("#", "ff")


def border2pyxlborder(nsb):
    if nsb.outline:
        nsb.outline.color: Color
        argb = color2argb(nsb.outline.color)
        s = PyxlSide(style=nsb.outline.style,
                     color=argb
                     )
        return PyxlBorder(left=s,
                          right=s,
                          top=s,
                          bottom=s,
                          )
    else:
        left = None
        if nsb.left:
            left = PyxlSide(style=nsb.left.style,
                            color=color2argb(nsb.left.color)
                            )
        right = None
        if nsb.left:
            right = PyxlSide(style=nsb.right.style,
                             color=color2argb(nsb.right.color)
                             )
        top = None
        if nsb.left:
            top = PyxlSide(style=nsb.top.style,
                           color=color2argb(nsb.top.color)
                           )
        bottom = None
        if nsb.left:
            bottom = PyxlSide(style=nsb.bottom.style,
                              color=color2argb(nsb.bottom.color)
                              )
        return PyxlBorder(left=left,
                          right=right,
                          top=top,
                          bottom=bottom)


def pf2pyxlpf(pf):
    return PyxlPatternFill(fill_type=pf.fill_type,
                           start_color=color2argb(pf.start_color),
                           end_color=color2argb(pf.end_color)
                           )


def font2pyxlfont(f: Font):
    return PyxlFont(bold=f.bold,
                    color=color2argb(f.color),
                    name=f.face
                    )

def add_named_styles(wbcwb, opwb):
    for ns in wbcwb.named_styles:
        the_style = PyxlNamedStyle(name=ns.name)
        if ns.font:
            the_style.font = font2pyxlfont(ns.font)
        if ns.border:
            the_style.border = border2pyxlborder(ns.border)
            opwb.add_named_style(the_style)
        if ns.pattern_fill:
            the_style.fill = pf2pyxlpf(ns.pattern_fill)
        opwb.add_named_style(the_style)

# opwb : openpxl workbook
# opsh : openpxl sheet
# wbcwb : workbook compiler Workbook
# wbcsh : workbook compiler Sheet
def render(wbcwb: Workbook):
    opwb = pxl.Workbook()
    add_named_styles(wbcwb, opwb)

    for wbcsh in wbcwb.sheets:
        opsh = opwb.create_sheet(wbcsh.name)
        for r in wbcsh.ranges:
            add_named_range(opwb, opsh, wbcsh, r)
        for r in wbcsh.ranges:
            fill_range_values(wbcwb, opsh, r)

    # If there is more than one sheet, remove the default.
    if len(opwb.sheetnames) > 1:
        opwb.remove(opwb["Sheet"])
    return opwb
